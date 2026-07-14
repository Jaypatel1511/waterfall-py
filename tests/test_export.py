"""Export: native DataFrame, JSON, and Excel (openpyxl)."""
import json
import os
from datetime import date

import pytest

from waterfall import run
from waterfall.data.schema import Deal, Tranche, ReserveConfig
from waterfall.report import export
from waterfall.report.schema import STANDARD_DISCLAIMER


def _deal():
    senior = Tranche(name="Term A", tranche_type="senior", principal=10_000_000.0,
                     coupon=0.06, amort_type="bullet", term_periods=4)
    equity = Tranche(name="Equity", tranche_type="equity", principal=3_000_000.0)
    return Deal(
        deal_close_date=date(2024, 1, 1), operations_start_date=date(2024, 1, 1),
        period_frequency="Q", deal_type="PF", tranches=[senior, equity],
        cfads_stream=[900_000.0] * 4, data_currency="USD", reporting_basis="calendar",
        reserves=[ReserveConfig(reserve_type="DSRA", opening_balance=300_000.0,
                                target_amount=300_000.0)],
        name="Demo Deal",
    )


def test_period_dataframe_shape_and_columns():
    df = export.to_period_dataframe(run(_deal()))
    assert len(df) == 4
    for col in ("period_index", "period_end_date", "cfads", "sweep_amount",
                "equity_distribution", "dscr", "covenant_status"):
        assert col in df.columns


def test_json_roundtrip_contains_disclaimer():
    payload = export.to_json(run(_deal()))
    data = json.loads(payload)
    assert data["disclaimer"] == STANDARD_DISCLAIMER
    assert len(data["periods"]) == 4
    assert data["deal_name"] == "Demo Deal"


def test_excel_written_and_reopenable(tmp_path):
    from openpyxl import load_workbook
    out = os.path.join(tmp_path, "deal.xlsx")
    export.to_excel(run(_deal()), out)
    assert os.path.isfile(out)
    wb = load_workbook(out)
    assert "period_table" in wb.sheetnames
    assert "audit_log" in wb.sheetnames
    ws = wb["period_table"]
    assert ws.max_row == 5   # header + 4 periods
