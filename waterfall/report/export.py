"""Export a DealResult to a DataFrame, Excel (openpyxl), JSON, or rendered text.

The rendered report opens with the standard disclaimer and the limitations, and
labels every breach as a mechanical test result — never a legal conclusion.
"""
import json

from waterfall.report.schema import DealResult, STANDARD_DISCLAIMER


def to_period_dataframe(result: DealResult):
    """The period table as a pandas DataFrame (methodology Output Schema columns)."""
    import pandas as pd

    rows = []
    for p in result.periods:
        rows.append({
            "period_index": p.period_index,
            "period_end_date": p.period_end_date,
            "cfads": p.cfads,
            "fees": p.fees,
            "facility_draws": p.facility_draws,
            "equity_contributions": p.equity_contributions,
            "debt_service_by_tranche": p.debt_service_by_tranche,
            "principal_by_tranche": p.principal_by_tranche,
            "interest_by_tranche": p.interest_by_tranche,
            "reserve_balances": p.reserve_balances,
            "reserve_draws": p.reserve_draws,
            "reserve_releases": p.reserve_releases,
            "proceeds_applied": p.proceeds_applied,
            "sweep_amount": p.sweep_amount,
            "junior_uses": p.junior_uses,
            "equity_distribution": p.equity_distribution,
            "covenant_status": p.covenant_status,
            "dscr": p.dscr,
        })
    return pd.DataFrame(rows)


def to_json(result: DealResult) -> str:
    """Serialize the result to a JSON string."""
    return json.dumps(_result_dict(result), default=str, indent=2)


def _result_dict(result: DealResult) -> dict:
    return {
        "deal_name": result.deal_name,
        "disclaimer": result.disclaimer,
        "limitations": result.limitations,
        "interpretation": result.interpretation,
        "source_use": result.source_use,
        "tranche_summary": result.tranche_summary,
        "periods": [_period_dict(p) for p in result.periods],
        "audit_log": [
            {"period_index": r.period_index, "event": r.event, "rationale": r.rationale,
             "mechanical_test_result": r.mechanical_test_result}
            for r in result.audit_log
        ],
    }


def _period_dict(p) -> dict:
    return {
        "period_index": p.period_index,
        "period_end_date": str(p.period_end_date),
        "cfads": p.cfads,
        "fees": p.fees,
        "facility_draws": p.facility_draws,
        "equity_contributions": p.equity_contributions,
        "debt_service_by_tranche": p.debt_service_by_tranche,
        "principal_by_tranche": p.principal_by_tranche,
        "interest_by_tranche": p.interest_by_tranche,
        "reserve_balances": p.reserve_balances,
        "reserve_draws": p.reserve_draws,
        "reserve_releases": p.reserve_releases,
        "proceeds_applied": p.proceeds_applied,
        "sweep_amount": p.sweep_amount,
        "junior_uses": p.junior_uses,
        "equity_distribution": p.equity_distribution,
        "covenant_status": p.covenant_status,
        "dscr": p.dscr,
    }


def to_excel(result: DealResult, path: str) -> str:
    """Write the result to an .xlsx workbook (period table, tranche summary,
    covenant status, source/use, audit log). Returns the path written."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "period_table"
    headers = ["period_index", "period_end_date", "cfads", "fees", "facility_draws",
               "equity_contributions", "reserve_draws", "reserve_releases",
               "proceeds_applied", "sweep_amount", "equity_distribution", "dscr"]
    ws.append(headers)
    for p in result.periods:
        ws.append([p.period_index, str(p.period_end_date), p.cfads, p.fees,
                   p.facility_draws, p.equity_contributions, p.reserve_draws,
                   p.reserve_releases, p.proceeds_applied, p.sweep_amount,
                   p.equity_distribution, p.dscr])

    ws2 = wb.create_sheet("tranche_summary")
    ws2.append(["tranche", "type", "original_principal", "ending_balance"])
    for row in result.tranche_summary:
        ws2.append([row["tranche"], row["type"], row["original_principal"],
                    row["ending_balance"]])

    ws3 = wb.create_sheet("source_use")
    for k, v in result.source_use.items():
        ws3.append([k, v])

    ws4 = wb.create_sheet("audit_log")
    ws4.append(["period_index", "event", "rationale", "mechanical_test_result"])
    for r in result.audit_log:
        ws4.append([r.period_index, r.event, r.rationale, r.mechanical_test_result])

    ws5 = wb.create_sheet("disclaimer")
    ws5.append([STANDARD_DISCLAIMER])
    for lim in result.limitations:
        ws5.append([lim])

    wb.save(path)
    return path


def render_text(result: DealResult) -> str:
    """A plain-text report. Opens with the disclaimer; labels breaches as
    mechanical test results (never legal conclusions)."""
    lines = []
    lines.append(f"Waterfall report — {result.deal_name}")
    lines.append("=" * 72)
    lines.append(STANDARD_DISCLAIMER)
    lines.append("")
    lines.append("Limitations:")
    for lim in result.limitations:
        lines.append(f"  - {lim}")
    lines.append("")
    lines.append("Period table:")
    header = f"{'per':>3} {'end':>10} {'cfads':>12} {'sweep':>12} {'equity_dist':>12} {'dscr':>8}"
    lines.append(header)
    for p in result.periods:
        dscr = "n/a" if p.dscr != p.dscr else f"{p.dscr:.2f}"
        lines.append(f"{p.period_index:>3} {str(p.period_end_date):>10} "
                     f"{p.cfads:>12,.0f} {p.sweep_amount:>12,.0f} "
                     f"{p.equity_distribution:>12,.0f} {dscr:>8}")
    lines.append("")
    breaches = [r for r in result.audit_log if r.mechanical_test_result]
    if breaches:
        lines.append("Breach flags (mechanical test results):")
        for r in breaches:
            lines.append(f"  period {r.period_index}: {r.rationale}")
    lines.append("")
    lines.append(result.interpretation)
    return "\n".join(lines)
